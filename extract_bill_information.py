#!/usr/bin/env python
"""
Extract bill information from PDF files and generate a report.
"""

from datetime import date
from email.mime import text
import sys

try:
    import coloredlogs
    import csv
    import locale
    import logging
    import os
    import pdfplumber
    import re
    import sys
    import yaml
    from argparse import ArgumentParser, RawTextHelpFormatter
    from datetime import datetime
    from dotenv import load_dotenv
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from os.path import basename, join, dirname, isfile, isdir
    from pprint import pprint
    from pathlib import Path
    import textfsm
    import gspread
    import pandas as pd
    from gspread_dataframe import set_with_dataframe, get_as_dataframe
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    import sqlite3
    import json
except ModuleNotFoundError as e:
    print(f"{e}. Did you load your environment?")
    sys.exit(1)

load_dotenv()
logger = logging.getLogger()


def main():
    with open(join(dirname(__file__), "defaults.yaml"), "r", encoding="utf-8") as f:
        defaults = yaml.safe_load(f)
    args = parse_command_line(defaults)
    args.defaults = defaults
    try:
        # locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')  # On Windows
        locale.setlocale(locale.LC_NUMERIC, args.locale)
        locale.setlocale(locale.LC_TIME, args.locale)
        for verbose_package in [
            "pdfminer.pdfpage",
            "pdfminer.pdfdocument",
            "pdfminer.psparser",
            "pdfminer.pdfinterp",
            "pdfminer.cmapdb",
            "pdfminer.pdfparser",
            "googleapiclient.discovery_cache",
        ]:
            verbose_logger = logging.getLogger(verbose_package)
            verbose_logger.setLevel(level=logging.ERROR)
        if args.debug:
            coloredlogs.install(level=logging.DEBUG, logger=logger)
        elif args.quiet:
            coloredlogs.install(level=logging.ERROR, logger=logger)
        else:
            coloredlogs.install(level=logging.INFO, logger=logger)
        return args.func(args)
    except Exception as e:
        logging.error(e, exc_info=True)
        return 1


def parse_command_line(defaults):
    parser = ArgumentParser(
        prog="extract_bill_information",
        description=__doc__,
        formatter_class=RawTextHelpFormatter,
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Run the program in debug. Default to 'False'.",
        required=False,
        default=False,
    )
    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Run the program in quiet mode. Default to 'False'.",
        required=False,
        default=False,
    )
    parser.add_argument(
        "--locale",
        help=f"The locale to use when reading the bills. By default '{defaults['locale']}'.",
        required=False,
        default=defaults["locale"],
    )
    parser.add_argument(
        "--use-cache",
        action="store_true",
        help=f"Use a cache of already processed bills. Default to True.",
        required=False,
        default=True,
    )
    parser.add_argument(
        "--no-use-cache",
        dest="use_cache",
        action="store_false",
        help="Do not use the cache.",
        required=False,
    )
    parser.add_argument(
        "--bill-input",
        help=f"The path where the bills are located. Either a directory where all bills will be processed or a file for just one bill. By default '{defaults['bill-input']}'.",
        required=False,
        default=defaults["bill-input"],
    )
    parser.add_argument(
        "--bill-input-excludes",
        type=lambda s: s.split(","),
        nargs="+",
        help=f"A comma separated list of paths to exclude. By default '{','.join(defaults['bill-input-excludes'])}'.",
        required=False,
        default=",".join(defaults["bill-input-excludes"]),
    )
    parser.add_argument(
        "--include-loads",
        action="store_true",
        help=f"Wether to add the load in the Workbook. By default '{defaults['include-loads']}'.",
        required=False,
        default=defaults["include-loads"],
    )
    parser.add_argument(
        "--load-input",
        help=f"The path where the laod CSV are located. By default '{defaults['load-input']}'.",
        required=False,
        default=defaults["load-input"],
    )
    parser.add_argument(
        "--load-input-excludes",
        type=lambda s: s.split(","),
        nargs="+",
        help=f"A comma separated list of paths to exclude. By default '{','.join(defaults['load-input-excludes'])}'.",
        required=False,
        default=",".join(defaults["load-input-excludes"]),
    )
    parser.add_argument(
        "--limit",
        type=int,
        help=f"Limit the number of files processed. By default '{defaults['limit']}' (-1 is no limit).",
        required=False,
        default=defaults["limit"],
    )
    parser.add_argument(
        "--dump",
        action="store_true",
        help="Just dump the text of the bills to stdout. Useful for debugging. Default to 'False'.",
        required=False,
        default=False,
    )
    parser.add_argument(
        "--dump-prefix",
        help=f"The prefix for the dump files. By default '{defaults['dump-prefix']}'.",
        required=False,
        default=defaults["dump-prefix"],
    )
    parser.add_argument(
        "--refresh-workbook",
        action="store_true",
        help=f"Refresh the Workbook file to upload to Google Sheets. Default to True.",
        required=False,
        default=True,
    )
    parser.add_argument(
        "--no-refresh-workbook",
        dest="refresh_workbook",
        action="store_false",
        help="Use a previously generated Workbook file to upload to Google Sheets.",
        required=False,
    )
    parser.add_argument(
        "--workbook",
        help=f"The name of the Workbook. By default '{defaults['workbook']}'.",
        required=False,
        default=defaults["workbook"],
    )
    parser.add_argument(
        "--no-trim-workbook",
        action="store_true",
        help=f"Trim unwanted columns from the Workbook. Useful for debugging. Default to 'False'.",
        required=False,
        default=False,
    )
    parser.add_argument(
        "--upload",
        action="store_true",
        help=f"Upload the extracted bill information to Google Sheets. Default to False.",
        required=False,
        default=True,
    )
    parser.add_argument(
        "--no-upload",
        dest="upload",
        action="store_false",
        help="Do not upload the extracted bill information to Google Sheets.",
        required=False,
    )
    parser.add_argument(
        "--gsheet-id",
        help=f"The ID of the Google Sheet to upload to. By default '{defaults['gsheet-id']}'.",
        required=False,
        default=defaults["gsheet-id"],
    )
    parser.add_argument(
        "--credentials",
        help=f"The service-account.json file to authenticate. By default '{defaults['credentials']}'.",
        required=False,
        default=defaults["credentials"],
    )
    parser.add_argument(
        "--incremental",
        action="store_true",
        help=f"Only update new bills in Google Sheets. Default to True.",
        required=False,
        default=True,
    )
    parser.add_argument(
        "--no-incremental",
        dest="incremental",
        action="store_false",
        help="Overwrite any existing Google Sheets.",
        required=False,
    )
    parser.set_defaults(func=extract_bill_information)
    return parser.parse_args()


def extract_bill_information(args):
    """
    Extracts and processes bill information from PDF files in a specified input directory.

    This function scans the given input directory for PDF files, extracts bill information using
    a dispatcher, sanitizes the extracted data, and organizes it by CUPS and bill ID. It also
    supports limiting the number of files processed and generates a final bill report.

    Args:
        args: An object containing the following attributes:
            - input_dir (str): Path to the directory containing PDF files.
            - defaults (dict): Default configuration, including 'dispatchers' for extraction.
            - limit (int): Maximum number of files to process. If 0 or less, all files are processed.

    Returns:
        int: 1 if an error occurs (e.g., input directory does not exist or no files found), otherwise None.
    """
    if args.refresh_workbook:
        if not isdir(args.bill_input) and not isfile(args.bill_input):
            logger.error(
                f"Input directory '{args.bill_input}' does not exist.")
            return 1

        bill_files, load_files = _list_files_to_process(args)
        if not bill_files:
            logger.error(f"No bills found in '{args.bill_input}'.")
            return 1

        if args.dump:
            _dump_bills(args.dump_prefix, bills)
            return 0

        bills = _extract_bills(args, bill_files)
        loads = {}
        if args.include_loads:
            loads = _extract_loads(args, load_files)
        wb = _generate_workbook(args, bills, loads)
    else:
        if not isfile(args.workbook):
            logger.error(f"Workbook '{args.workbook}' does not exist.")
            return 1
        logger.info(
            f"Loading previously generated workbook '{args.workbook}' ...")
        wb = load_workbook(args.workbook, data_only=True)

    if args.upload:
        if not args.gsheet_id:
            logger.error(
                f"When uploading to GoogleSheet you must set the 'gsheet-id' parameter."
            )
            return 1
        bill_id_column = args.defaults["column_labels"]["bill_id"]
        _upload_report(wb, bill_id_column, args.credentials,
                       args.gsheet_id, args.incremental)

    return 0


def _list_files_to_process(args):
    """Returns the list of bills (and loads) to be processed.

    Args:
        args (namespace): the command line arguments

    Returns:
        tuple: the list of bills and the list of loads.
    """
    bills = loads = []

    bills = _list_files_with_extension(
        args.bill_input, args.bill_input_excludes, "pdf")
    if bills:
        logger.info(f"Found {len(bills)} bills in '{args.bill_input}'.")

    if args.include_loads:
        loads = _list_files_with_extension(
            args.load_input, args.load_input_excludes, "csv")
        if loads:
            logger.info(f"Found {len(loads)} loads in '{args.bill_input}'.")

    if args.limit > 0 and (bills or loads):
        bills = bills[: args.limit]
        loads = loads[: args.limit]
        logger.warning(
            f"Limiting to {args.limit} bills as per the --limit argument.")

    return bills, loads


def _list_files_with_extension(input, excludes, extension):
    """Return the list of file with the specific extension

    Args:
        input (str): the input, either a file or a directory
        excludes (list): a list of  regexp expression to exclude when walking from the input directory
        extension (str): the file extension to look for

    Returns:
        list: the list of included files
    """
    input_excludes = [re.compile(exclude)for exclude in excludes]
    if isfile(input):
        files = [input]
    else:
        files = [
            full_path
            for dp, dn, filenames in os.walk(input)
            for filename in filenames
            for full_path in [join(dp, filename)]
            if filename.lower().endswith(f".{extension}")
            and not any(regex.search(full_path) for regex in input_excludes)
        ]
    return files


def _dump_bills(dump_prefix, bills):
    """Dump extracted text from a list of bills to a specific folder

    Args:
        dump_prefix (str): the output prefix
        bills (list): the list of pdf files to extract and dump
    """
    for bill in bills:
        logger.info(bill)
        with open(
            f"{dump_prefix}{basename(bill).replace(".pdf", "")}.txt", "w", encoding="utf-8"
        ) as f:
            with pdfplumber.open(bill) as pdf:
                for page in pdf.pages:
                    f.write(page.extract_text())
    logger.info(
        f"Dumped {len(bills)} bills to {dirname(dump_prefix)}. Exiting."
    )


def _extract_bills(args, files: list):
    """Process all the input bills and construct a dict by CUPS/BILL_ID with all the relevant information

    Args:
        args (namespace): the input arguments
        files (list): the list of files to process

    Returns:
        dict: a dict of CUPS/BILL_ID with all the extracted information
    """
    known_bills = {}
    conn = None
    if args.use_cache:
        conn, known_bills = _load_cache(args.defaults['cache-file'])
    bills = {}
    try:
        for file in sorted(files):
            if args.use_cache and basename(file) in known_bills:
                logger.info("Using cache for bill '%s' ...",
                            _readable_path(file))
                bill_info = known_bills[basename(file)]
            else:
                bill_info = _extract_dispatcher(
                    args.defaults["dispatchers"], file)
                if not bill_info:
                    continue
                bill_info = _sanitize_bill(bill_info)
                if not bill_info:
                    logger.debug(f"Skipping {file} ...")
                    continue
                bill_info["file"] = basename(file)
                if args.use_cache:
                    _mark_as_processed(conn, bill_info)
            cups = bill_info["cups"]
            if cups not in bills:
                bills[cups] = []
            bills[cups].append(bill_info)
    except KeyboardInterrupt:
        logger.info(f"Interrupted by user. Cleaning up. Please wait.")
    finally:
        if conn:
            conn.close()

    return bills


def _extract_dispatcher(dispatchers: dict, file: str):
    """
    Extracts bill information by selecting and invoking the
    appropriate extractor function based on the content of
    the first page of a PDF bill.

    Args:
        dispatchers (dict): A dictionary mapping dispatcher names (str) to extractor function names (str).
        file (str): The path to the PDF file to be processed.

    Returns:
        Any: The result of the extractor function if a matching dispatcher is found; otherwise, None.

    Raises:
        ValueError: If the specified extractor function is not found or is not callable.

    """
    logger.info("Extracting information from bill '%s' ...",
                _readable_path(file))
    with pdfplumber.open(file) as pdf:
        first_page = pdf.pages[0].extract_text()
        found_extractor = False
        for dispatcher, extractor in dispatchers.items():
            if extractor not in globals() or not callable(globals()[extractor]):
                raise ValueError(
                    f"Extractor '{extractor}' not found or not callable.")
            if dispatcher in first_page:
                logger.debug(
                    f"Detected '{dispatcher}' bill. Using {extractor} to extract information ..."
                )
                bill_info = globals()[extractor](pdf)
                return bill_info
        if not found_extractor:
            logger.error(
                f"Could not detect the type of bill for '{file}'. Skipping")
            return None, None, None


def extract_plenitude_bill(pdf):
    """Extractor for plenitude bills

    Args:
        pdf (PDFfile): The corresponding pdf file

    Returns: 
        dict: A dictionary containing the extracted bill information.
    """
    if not hasattr(extract_plenitude_bill, "re_table"):
        extract_plenitude_bill.re_table = None
    numeric_cols = [
        "billed_amount",
        "billed_energy",
        "billed_power",
        "P1",
        "P2",
        "P3",
        "P4",
        "P5",
        "P6",
        "CP1",
        "CP2",
        "CP3",
        "CP4",
        "CP5",
        "CP6",
    ]
    with open(
        join(dirname(__file__), "assets/templates/es.plenitude.textfsm"),
        "r",
        encoding="utf-8",
    ) as template:
        extract_plenitude_bill.re_table = textfsm.TextFSM(template)
    df = _extract_bill(pdf, extract_plenitude_bill.re_table, numeric_cols)
    return df.iloc[0].to_dict()


def extract_nufri_bill(pdf):
    """Extractor for Nufri bills

    Args:
        pdf (PDFfile): The corresponding pdf file

    Returns:
        dict: A dictionary containing the extracted bill information.
    """
    if not hasattr(extract_nufri_bill, "re_table"):
        extract_nufri_bill.re_table = None

    numeric_cols = [
        "billed_amount",
        "billed_energy",
        "billed_power",
        "P1",
        "P2",
        "P3",
        "P4",
        "P5",
        "P6",
    ]
    with open(
        join(dirname(__file__), "assets/templates/es.nufri.textfsm"),
        "r",
        encoding="utf-8",
    ) as template:
        extract_nufri_bill.re_table = textfsm.TextFSM(template)
    df = _extract_bill(pdf, extract_nufri_bill.re_table, numeric_cols)
    matches = re.findall(r"(P[1-6])\s+([\d,]+)\s+kW",
                         df["nufri_contracted_power"][0])
    if matches:
        for i in range(1, 7):
            df[f"CP{i}"] = 0.0
        for k, v in dict(matches).items():
            df[f"C{k}"] = locale.atof(v)
    return df.iloc[0].to_dict()


def extract_te_bill(pdf):
    """Extractor for Total Energie bills

    Args:
        pdf (PDFfile): The corresponding pdf file

    Returns:
        dict: A dictionary containing the extracted bill information.
    """
    if not hasattr(extract_te_bill, "re_table"):
        extract_te_bill.re_table = None

    numeric_cols = [
        "billed_amount",
        "te_power_access_P1",
        "te_power_access_P2",
        "te_power_access_P3",
        "te_power_access_P4",
        "te_power_access_P5",
        "te_power_access_P6",
        "te_power_P1",
        "te_power_P2",
        "te_power_P3",
        "te_power_P4",
        "te_power_P5",
        "te_power_P6",
        "te_power_charge_P1",
        "te_power_charge_P2",
        "te_power_charge_P3",
        "te_power_charge_P4",
        "te_power_charge_P5",
        "te_power_charge_P6",
        "te_energy_access_P1",
        "te_energy_access_P2",
        "te_energy_access_P3",
        "te_energy_access_P4",
        "te_energy_access_P5",
        "te_energy_access_P6",
        "te_energy_P1",
        "te_energy_P2",
        "te_energy_P3",
        "te_energy_P4",
        "te_energy_P5",
        "te_energy_P6",
        "te_energy_charge_P1",
        "te_energy_charge_P2",
        "te_energy_charge_P3",
        "te_energy_charge_P4",
        "te_energy_charge_P5",
        "te_energy_charge_P6",
        "P1",
        "P2",
        "P3",
        "P4",
        "P5",
        "P6",
    ]

    with open(
        join(dirname(__file__), "assets/templates/es.totalenergies.textfsm"),
        "r",
        encoding="utf-8",
    ) as template:
        extract_te_bill.re_table = textfsm.TextFSM(template)
    df = _extract_bill(pdf, extract_te_bill.re_table, numeric_cols)
    # Do some summing
    df["billed_energy"] = df[
        [
            f"{prefix}{i}"
            for i in range(1, 7)
            for prefix in ("te_energy_access_P", "te_energy_P", "te_energy_charge_P")
        ]
    ].sum(axis=1)
    df["billed_power"] = df[
        [
            f"{prefix}{i}"
            for i in range(1, 7)
            for prefix in ("te_power_access_P", "te_power_P", "te_power_charge_P")
        ]
    ].sum(axis=1)
    # unfortunately te_contracted_power is not easily extractible
    matches = re.findall(r"(P[1-6])\s+([\d,]+)", df["te_contracted_power"][0])
    if matches:
        for i in range(1, 7):
            df[f"CP{i}"] = 0.0
        for k, v in dict(matches).items():
            df[f"C{k}"] = locale.atof(v)
    return df.iloc[0].to_dict()


def extract_endesa_bill(pdf):
    """Extractor for Endesa bills

    Args:
        pdf (PDFfile): The corresponding pdf file

    Returns:
        dict: A dictionary containing the extracted bill information.
    """
    if not hasattr(extract_endesa_bill, "re_table"):
        extract_endesa_bill.re_table = None
    numeric_cols = [
        "billed_amount",
        "billed_power",
        "billed_energy",
        "P1",
        "P2",
        "P3",
        "P4",
        "P5",
        "P6",
        "CP1",
        "CP2",
        "CP3",
        "CP4",
        "CP5",
        "CP6",
    ]

    with open(
        join(dirname(__file__), "assets/templates/es.endesa.textfsm"),
        "r",
        encoding="utf-8",
    ) as template:
        extract_endesa_bill.re_table = textfsm.TextFSM(template)
    df = _extract_bill(pdf, extract_endesa_bill.re_table, numeric_cols)
    # I need to adjust the billing period end by -1 otherwise I
    # overshoot the contracted power calculation. Except when it's
    # just one day
    # if df["billing_period_start"][0] != df["billing_period_end"][0]:
    # XXX: check that is correct
    # df["billing_period_end"][0] = df["billing_period_end"][0] - 1
    # pass
    # matches = re.findall(
    #     r"\s*(punta|punta-llano|valle)\s*([\d,]+)\s*kW;?",
    #     df["endesa_contracted_power"][0],
    # )
    # if matches:
    #     for i in range(1, 7):
    #         df[f"CP{i}"] = 0.0
    #     for k, v in dict(matches).items():
    #         if k == "punta" or k == "punta-llano":
    #             df[f"CP1"] = locale.atof(v)
    #         elif k == "valle":
    #             df[f"CP3"] = locale.atof(v)
    #         else:
    #             logger.warning(f"Unknown contracted power {k}")
    return df.iloc[0].to_dict()


def extract_qener_bill(pdf):
    """Extractor for Qener bills

    Args:
        pdf (PDFfile): The corresponding pdf file

    Returns:
        dict: A dictionary containing the extracted bill information.
    """
    if not hasattr(extract_qener_bill, "re_table"):
        extract_qener_bill.re_table = None
    numeric_cols = [
        "billed_amount",
        "billed_energy",
        "billed_power",
        "P1",
        "P2",
        "P3",
        "P4",
        "P5",
        "P6",
        "CP1",
        "CP2",
        "CP3",
        "CP4",
        "CP5",
        "CP6",
    ]
    with open(
        join(dirname(__file__), "assets/templates/es.qener.textfsm"),
        "r",
        encoding="utf-8",
    ) as template:
        extract_qener_bill.re_table = textfsm.TextFSM(template)
    df = _extract_bill(pdf, extract_qener_bill.re_table, numeric_cols)
    return df.iloc[0].to_dict()


def _extract_bill(pdf, re_table, numeric_columns):
    pages = [page.extract_text() for page in pdf.pages]
    headers = re_table.header
    data = re_table.ParseText("\n".join(pages))
    df = pd.DataFrame(data, columns=headers)
    # Convert to float and fill with 0
    df[numeric_columns] = (
        df[numeric_columns]
        .apply(
            lambda serie: pd.to_numeric(
                serie.astype(str)
                .str.replace(".", "", regex=False)
                .str.replace(",", ".", regex=False),
                errors="coerce",
            )
        )
        .fillna(0)
    )
    return df


def _sanitize_bill(data: dict):
    """
    Sanitize the bill data by converting strings to appropriate types and formatting.
    """
    is_sane, missing_keys = _is_bill_sane(data)
    if not is_sane:
        logger.error(
            f"Missing keys for CUPS {data['cups']} and bill {data['bill_id']}: {missing_keys}"
        )
        return None

    if not data["cups"].endswith("0F"):
        # For some reason all CUPS ends with 0F except for Qener bills...
        cups = f"{data['cups']}0F"
        data["cups"] = cups

    if not data["billing_date"]:
        logger.error(
            f"Invalid billing date for CUPS {data['cups']} and bill {data['bill_id']}"
        )
        return None

    if not data["billing_period_start"] or not data["billing_period_end"]:
        logger.error(
            f"Invalid billing period for CUPS {data['cups']} and bill {data['bill_id']}"
        )
        return None

    if "billed_power" not in data:
        logger.error(
            f"Invalid billed power capacity (not in the expected format '+/-0.000,00 €') for CUPS {data['cups']} and bill {data['bill_id']}."
        )
        return None

    if "billed_energy" not in data:
        logger.error(
            f"Invalid billed energy consumed (not in the expected format '+/-0.000,00 €') for CUPS {data['cups']} and bill {data['bill_id']}."
        )
        return None

    if "billed_amount" not in data:
        logger.error(
            f"Invalid billed amount (not in the expected format '+/-0.000,00 €') for CUPS {data['cups']} and bill {data['bill_id']}."
        )
        return None

    for power_type in ["P1", "P2", "P3", "P4", "P5", "P6"]:
        if power_type not in data and power_type in ["P1", "P2", "P3"]:
            logger.error(
                f"Mandatory '{power_type}' comsumption has not been extracted from CUPS {data['cups']} and bill {data['bill_id']}."
            )
            return None
        if (
            power_type in data
            and data[power_type] is not None
            and type(data[power_type]) is str
        ):  # P4, P5, P6
            data[power_type] = locale.atof(data[power_type].strip())

    return data


def _is_bill_sane(data: dict):
    # check that all the keys have a value
    is_sane = True
    missing_keys = []
    for key, value in data.items():
        if value is None or (isinstance(value, str) and not value.strip()):
            missing_keys.append(key)
            is_sane = False
    return is_sane, missing_keys


def _readable_path(path: str, max_len: int = 70) -> str:
    """Return a user friendly path to be displayed on screen

    Args:
        path (str): the complete file path
        max_len (int, optional): the maximum number of characters to be returned. Defaults to 70.

    Returns:
        str: the truncated path
    """
    path = Path(path)
    parts = path.parts

    # Return unchanged if already short enough
    if len(str(path)) <= max_len:
        return str(path)

    # Keep first element and as many trailing elements as possible
    for n in range(len(parts) - 1, 1, -1):
        shortened = str(Path(parts[0], "...", *parts[-n:]))
        if len(shortened) <= max_len:
            return shortened

    # Fallback
    return f"{parts[0]}.../{parts[-1]}"


def _extract_loads(args, files: list):
    loads = {}
    for file in files:
        with open(file, encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter=";")
            for row in reader:
                cups = row["CUPS"].strip()
                fecha = row["Fecha"].strip()
                hora = row["Hora"].strip()
                ae_kwh = row["AE_kWh"].strip()

                # Hora starts at 1, while datetime starts at 0
                dt_str = f"{fecha} {int(hora)-1}"
                try:
                    dt = datetime.strptime(dt_str, "%d/%m/%Y %H")
                except ValueError as e:
                    logger.error(
                        f"Could not parse datetime '{dt_str}' in file '{file}': {str(e)}"
                    )
                    continue

                # Convert AE_kWh to float
                try:
                    if ae_kwh != "":
                        ae_kwh_val = locale.atof(ae_kwh)
                    else:
                        ae_kwh_val = 0.0
                except ValueError as e:
                    logger.error(
                        f"Could not parse AE_kWh '{ae_kwh}' in file '{file}'")
                    continue

                # Insert into loads dict
                if cups not in loads:
                    loads[cups] = {}
                loads[cups][dt] = ae_kwh_val
    return loads


def _generate_workbook(args, bills: dict, loads: dict) -> Workbook:
    """
    Generates an Excel workbook report from provided bill information.

    For each CUPS (supply point) in the `bills` dictionary, a worksheet is created and populated
    with bill data. The worksheet names and column headers are determined by the `args.defaults`
    configuration. The resulting workbook is saved to the file path specified by `args.workbook`.

    Args:
        args: An object containing configuration options, including:
            - defaults['column_labels']: A dictionary mapping column keys to their display labels.
            - defaults['sheet_names']: A dictionary mapping CUPS to worksheet names.
            - workbook: The file path where the Excel workbook will be saved.
        bills (dict): A dictionary where each key is a CUPS identifier and each value is another
            dictionary containing bill information for that CUPS.
        loads (dict): A dictionary where each key is a CUPS identifier and each value is another
            dictionary containing load information for that CUPS.

    Returns:
        Workbook. The function generated workbook
    """
    wb = Workbook()

    column_keys = list(args.defaults["column_labels"].keys())
    column_headers = list(args.defaults["column_labels"].values())
    sheet_names = args.defaults["sheet_names"]

    ws = wb.active
    ws.title = "Loads"
    ws.append(["CUPS", "Fecha", "AE_kWh"])
    for cups, cups_loads in loads.items():
        for dt, load in cups_loads.items():
            ws.append([cups, dt, load])

    sheets = {}
    # Add a new worksheet for each CUPS
    for cups, bill_infos in bills.items():
        logger.info(
            f"Adding worksheet for CUPS '{cups}' with {len(bill_infos)} bills ..."
        )
        ws = wb.create_sheet(title=sheet_names.get(cups, cups))
        sheets[ws.title] = ws

        # Enrich the report
        if args.no_trim_workbook:
            df = pd.DataFrame(bill_infos)
        else:
            df = pd.DataFrame(bill_infos, columns=column_keys)
        df["gross_amount"] = df["billed_power"] + df["billed_energy"]
        df = df.rename(columns=args.defaults["column_labels"])

        # And write it in the worksheet
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

    ordered_sheets = []
    for title in args.defaults["sheet_names"].values():
        if title in sheets:
            ordered_sheets.append(sheets[title])
    if len(ordered_sheets) == 0:
        ordered_sheets.extend(list(sheets.values()))
    wb._sheets = [wb.worksheets[0]] + ordered_sheets

    # Save the workbook
    wb.save(args.workbook)
    return wb


def _upload_report(workbook: Workbook, bill_id_column: str, cred_file: str, gsheet_id: str, incremental: bool):
    logger.info(
        msg=f"Uploading report to Google Sheets with ID '{gsheet_id}' ...")
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    creds = Credentials.from_service_account_file(cred_file, scopes=scopes)
    g_spreadsheet = gspread.authorize(creds).open_by_key(gsheet_id)

    for l_worksheet in workbook.worksheets:
        if incremental and l_worksheet.title != 'Loads':
            _update_worksheet_incremental(
                l_worksheet, g_spreadsheet, bill_id_column)
        else:
            _update_worksheet_overwrite(l_worksheet, g_spreadsheet)

    drive = build("drive", "v3", credentials=creds)
    sheet = (
        drive.files()
        .get(
            fileId=gsheet_id,
            fields="id, kind, mimeType, name, size, modifiedTime, webViewLink",
        )
        .execute()
    )
    logger.info(f"Report '{sheet['name']}' saved to Google Drive ...")


def _update_worksheet_incremental(l_worksheet, g_spreadsheet, bill_id_column: str):
    """Update a specific google worksheet with the values found in the openpyxl worksheet

    Args:
        l_worksheet (worksheet): the openpyxl worksheet
        g_spreadsheet (GoogleWorkheet): the google worksheet to update
        bill_id_column (str): the column that represent the billd id (its localized form)
    """
    l_df = _get_df_from_worksheet(l_worksheet)
    l_title = l_worksheet.title

    try:
        g_worksheet = g_spreadsheet.worksheet(l_title)
    except gspread.exceptions.WorksheetNotFound:
        # Google worksheet does not exists ...
        logger.info(f"Creating new sheet '{l_title}' ...")
        g_worksheet = g_spreadsheet.add_worksheet(
            title=l_worksheet.title, rows=l_worksheet.max_row, cols=l_worksheet.max_column
        )
        _write_to_worksheet(l_df, g_worksheet)
        return

    g_df = get_as_dataframe(g_worksheet)
    if len(g_df) <= 1:
        # Google worksheet is empty ...
        logger.info(f"Overwriting empty sheet '{l_title}' ...")
        _write_to_worksheet(l_df, g_worksheet)
        return

    if bill_id_column not in g_df.columns:
        # Google worksheet does not contain the bill_id column
        logger.warning(
            f"Could not detect bill_id column '{bill_id_column}' in existing worksheet '{l_title}'. Overwriting Google worksheet."
        )
        g_worksheet.clear()
        _write_to_worksheet(l_df, g_worksheet)
        return

    def _extract_bill_id(value):
        if pd.isna(value): return value
        m = re.match(
            r'=HYPERLINK\(".*?";\s*"([^"]+)"\)',
            str(value)
        )
        # Either a plain bill id or an already transformed hyperlink
        return m.group(1) if m else str(value)

    g_bill_ids = g_df[bill_id_column].apply(_extract_bill_id)
    new_bills = l_df.loc[~l_df[bill_id_column].isin(g_bill_ids)]
    if new_bills.empty:
        logger.info(f"No new bill detected in worksheet '{l_title}' ...")
        return
    
    logger.info(f"Appending {len(new_bills)} new bills to worksheet '{l_title}' ...")
    g_worksheet.append_rows(
        new_bills.values.tolist(),
        value_input_option="USER_ENTERED"
    )


def _update_worksheet_overwrite(l_worksheet, g_spreadsheet):
    """Overwrite a google worksheet with a openpyxl worksheet

    Args:
        l_worksheet (worksheet): the openpyxl worksheet
        g_spreadsheet (Google worksheet): the Google worksheet to update
    """
    l_df = _get_df_from_worksheet(l_worksheet)
    l_title = l_worksheet.title

    try:
        g_worksheet = g_spreadsheet.worksheet(l_title)
        g_spreadsheet.del_worksheet(g_worksheet)
    except gspread.exceptions.WorksheetNotFound:
        pass
    rows, cols = l_df.shape
    g_worksheet = g_spreadsheet.add_worksheet(
        title=l_title, rows=rows, cols=cols)
    logger.info(f"Overwriting worksheet '{l_title}' ...")
    _write_to_worksheet(l_df, g_worksheet)


def _write_to_worksheet(l_df, g_worksheet):
    """Write a pandas dataframe to a Google worksheet

    Args:
        l_df (dataframe): the pandas dataframe
        g_worksheet (Google Worksheet): the Google worksheet to update
    """
    set_with_dataframe(g_worksheet, l_df, include_index=False,
                       include_column_header=True)


def _get_df_from_worksheet(l_worksheet):
    """Return a pandas dataframe from the Openpyxl worksheet

    Args:
        l_worksheet (worksheet): the openpyxl worksheet

    Returns:
        df: pandas dataframe
    """
    data = l_worksheet.values
    columns = next(data)
    l_df = pd.DataFrame(data, columns=columns)
    return l_df


def _load_cache(cache_file: Path):
    """load the cache from the given sqlite database

    Args:
        cache_file (Path): the path to the sqlite database

    Returns:
        tuple: the sqlite connection and a dict of known bill, with the filename as key
    """
    conn = sqlite3.connect(cache_file)
    conn.execute("""
    CREATE TABLE IF NOT EXISTS extracted_bills (
        file TEXT PRIMARY KEY,
        bill_info TEXT,
        exctracted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)
    conn.commit()
    cur = conn.execute("SELECT file, bill_info FROM extracted_bills")
    known_bills = {row[0]: json.loads(row[1]) for row in cur}
    return conn, known_bills


def _mark_as_processed(conn, bill_info: dict):
    """Store the bill in the cache

    Args:
        conn (sqlite3.connection): the sqlite connection
        bill_info (dict): the bill information
    """
    bill_info_ser = json.dumps(bill_info)
    conn.execute(
        "INSERT OR IGNORE INTO extracted_bills (file, bill_info) VALUES (?, ?)", (bill_info["file"], bill_info_ser, ))
    conn.commit()


if __name__ == "__main__":
    sys.exit(main())
